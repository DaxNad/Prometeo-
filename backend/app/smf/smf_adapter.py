from __future__ import annotations

import asyncio
import os
from pathlib import Path
import logging
import traceback
from openpyxl import Workbook
from typing import Any

from .smf_reader import SMFReader
from .smf_writer import SMFWriter
from .smf_updater import SMFUpdater
from .smf_schema import REQUIRED_SCHEMA


MASTER_NAME = "SuperMegaFile_Master.xlsx"


def _default_smf_dir() -> Path:
    env_dir = os.getenv("SMF_BASE_PATH", "").strip()
    if env_dir:
        return Path(env_dir)

    # Railway Runtime: prefer a consistent writable mount
    if os.getenv("RAILWAY") or os.getenv("RAILWAY_ENVIRONMENT") or os.getenv("RAILWAY_STATIC_URL"):
        return Path("/data/local_smf")

    return Path.home() / "Documents" / "local_smf"


class SMFAdapter:
    def __init__(self, base_path: Path | None = None) -> None:
        self.base_path = base_path or _default_smf_dir()
        self._agent_runtime = None
        self._bootstrap_if_missing()

    def master_path(self) -> Path:
        return self.base_path / MASTER_NAME

    def available(self) -> bool:
        return self.master_path().exists()

    def reader(self) -> SMFReader:
        return SMFReader(self.master_path())

    def writer(self) -> SMFWriter:
        return SMFWriter(self.master_path())

    def updater(self) -> SMFUpdater:
        return SMFUpdater(self.master_path())

    def info(self) -> dict:
        payload = {
            "base_path": str(self.base_path),
            "path": str(self.master_path()),
            "exists": self.available(),
        }

        self._agent_monitor(
            event_type="smf_info",
            severity="info",
            payload=payload,
        )

        return payload

    def structure(self) -> dict:
        reader = self.reader()
        payload = {
            "exists": reader.exists(),
            "sheets": reader.sheets(),
        }

        self._agent_monitor(
            event_type="smf_structure",
            severity="info",
            payload={
                "exists": payload["exists"],
                "sheets_count": len(payload["sheets"]),
                "sheet_names": payload["sheets"],
            },
        )

        return payload

    def validate_structure(self) -> dict:
        """Validate presence of required sheets and columns.

        Returns a dict with `ok`, and details about missing/extra parts.
        """
        import pandas as pd

        result = {
            "ok": False,
            "sheets_missing": [],
            "columns_missing": {},
            "columns_extra": {},
        }

        if not self.available():
            result["sheets_missing"] = list(REQUIRED_SCHEMA.keys())
            return result

        xls = pd.ExcelFile(self.master_path())
        available = set(xls.sheet_names)
        required = set(REQUIRED_SCHEMA.keys())
        missing_sheets = sorted(list(required - available))
        result["sheets_missing"] = missing_sheets

        for sheet in sorted(required & available):
            df = pd.read_excel(self.master_path(), sheet_name=sheet)
            have = set(map(str, list(df.columns)))
            need = set(REQUIRED_SCHEMA[sheet])
            missing_cols = sorted(list(need - have))
            extra_cols = sorted(list(have - need))
            if missing_cols:
                result.setdefault("columns_missing", {})[sheet] = missing_cols
            if extra_cols:
                result.setdefault("columns_extra", {})[sheet] = extra_cols

        result["ok"] = not result["sheets_missing"] and not result["columns_missing"]
        return result

    # --- internal ---------------------------------------------------------
    def _bootstrap_if_missing(self) -> None:
        """Ensure base path exists and create an empty SMF workbook if missing.

        This makes SMF endpoints resilient when SuperMegaFile is not present
        (first run, Railway fresh deploy, etc.).
        """
        self._bootstrap_attempted = True  # diagnostics
        self._bootstrap_error = None

        base = self.base_path
        master = self.master_path()
        try:
            logger = logging.getLogger(__name__)
            logger.debug("SMF bootstrap: base_path=%s master_path=%s", base, master)
            base.mkdir(parents=True, exist_ok=True)
            logger.debug(
                "SMF bootstrap env: base_exists=%s base_is_dir=%s master_exists_before=%s writable_check=%s",
                base.exists(), base.is_dir(), master.exists(), self._writable_check(base)
            )
            if master.exists():
                logger.debug("SMF bootstrap: master already present, skipping creation")
                return

            wb = Workbook()

            # Sheet: Codici
            ws_codes = wb.active
            ws_codes.title = "Codici"
            ws_codes.append(["Codice", "Descrizione"])

            # Sheet: Postazioni
            ws_stations = wb.create_sheet("Postazioni")
            ws_stations.append(["Postazione", "Note"])

            # Sheet: Pianificazione
            ws_plan = wb.create_sheet("Pianificazione")
            ws_plan.append([
                "ID ordine",
                "Cliente",
                "Codice",
                "Q.ta",
                "Data richiesta cliente",
                "Priorità",
                "Postazione assegnata",
                "Stato (da fare/in corso/finito)",
                "Progress %",
                "Semaforo scadenza",
                "Note",
            ])

            # Sheet: DiscrepanzeLog (for diagnostics)
            ws_log = wb.create_sheet("DiscrepanzeLog")
            ws_log.append([
                "Timestamp",
                "Sorgente",
                "Voce",
                "Valore atteso",
                "Valore trovato",
                "Azione eseguita",
                "Esito",
            ])

            wb.save(master)
            logger.info("SMF bootstrap: created new workbook at %s (exists_after=%s)", master, master.exists())
        except Exception:
            self._bootstrap_error = traceback.format_exc()
            # Avoid dumping full trace to stdout in production; keep it attached for diagnostics endpoint
            logging.getLogger(__name__).error("SMF bootstrap failed: %s", str(self._bootstrap_error).splitlines()[-1])

    def _writable_check(self, path: Path) -> bool:
        try:
            path.mkdir(parents=True, exist_ok=True)
            test = path / ".smf_write_test"
            with open(test, "w", encoding="utf-8") as f:
                f.write("ok")
            test.unlink(missing_ok=True)  # type: ignore[call-arg]
            return True
        except Exception:
            return False

    def preview(self, sheet: str | None = None, rows: int = 5) -> dict:
        reader = self.reader()
        payload = reader.preview(sheet=sheet, rows=rows)

        self._agent_monitor(
            event_type="smf_preview",
            severity="info",
            payload={
                "sheet": sheet or "AUTO",
                "rows_requested": rows,
                "ok": payload.get("ok"),
                "exists": payload.get("exists"),
            },
        )

        return payload

    def append_order(self, data: dict) -> dict:
        payload = self.writer().append_row("Pianificazione", data)

        self._agent_monitor(
            event_type="smf_append_order",
            severity="info",
            payload={
                "sheet": "Pianificazione",
                "input_keys": sorted(list(data.keys())),
                "ok": payload.get("ok"),
            },
        )

        return payload

    def update_order(self, order_id: str, updates: dict) -> dict:
        payload = self.updater().update_row(
            sheet="Pianificazione",
            id_column="ID ordine",
            id_value=order_id,
            updates=updates,
        )

        self._agent_monitor(
            event_type="smf_update_order",
            severity="info",
            payload={
                "sheet": "Pianificazione",
                "id_column": "ID ordine",
                "id_value": order_id,
                "update_keys": sorted(list(updates.keys())),
                "ok": payload.get("ok"),
            },
        )

        return payload

    def _agent_monitor(
        self,
        *,
        event_type: str,
        severity: str = "info",
        payload: dict[str, Any] | None = None,
    ) -> None:
        agent_runtime = self._get_agent_runtime()
        if agent_runtime is None:
            return
        try:
            asyncio.run(
                agent_runtime.analyze(
                    source="smf_adapter",
                    line_id="smf",
                    event_type=event_type,
                    severity=severity,
                    payload=payload or {},
                )
            )
        except RuntimeError:
            pass
        except Exception:
            pass

    def _get_agent_runtime(self):
        if self._agent_runtime is not None:
            return self._agent_runtime
        try:
            from app.agent_runtime.service import AgentRuntimeService
        except Exception:
            return None
        try:
            self._agent_runtime = AgentRuntimeService()
        except Exception:
            return None
        return self._agent_runtime
