from pathlib import Path
import warnings

import openpyxl

from app.smf.smf_adapter import SMFAdapter


def _build_empty_like_planning(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pianificazione"
    # Intestazioni canoniche, ma senza alcun valore: pandas può inferire dtype non-object
    ws.append(
        [
            "ID ordine",
            "Cliente",
            "Codice",
            "Q.ta",
            "Data richiesta cliente",
            "Priorità",
            "Postazione assegnata",
            "Stato (da fare/in corso/finito)",
            "Progress %",
            "Semaforo scadenza",
            "Note",
            # colonne alias presenti in altri test
            "order_id",
            "cliente",
            "codice",
            "qta",
            "postazione",
            "stato",
        ]
    )

    # Riga di esempio con molti None per indurre dtype float nelle colonne stringa
    ws.append([
        "ROW-1",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "ROW-1",
        None,
        None,
        None,
        None,
        None,
    ])

    # Aggiungiamo i fogli minimi richiesti dal writer/updater
    codes = wb.create_sheet("Codici")
    codes.append(["Codice", "Descrizione"])  # header

    stations = wb.create_sheet("Postazioni")
    stations.append(["Postazione", "Note"])  # header

    log = wb.create_sheet("DiscrepanzeLog")
    log.append(["Timestamp", "Sorgente", "Voce", "Valore atteso", "Valore trovato", "Azione eseguita", "Esito"])  # header

    wb.save(path)


def test_update_row_does_not_raise_futurewarning(tmp_path: Path):
    smf_dir = tmp_path / "smf_warn"
    smf_dir.mkdir()
    workbook_path = smf_dir / "SuperMegaFile_Master.xlsx"
    _build_empty_like_planning(workbook_path)

    adapter = SMFAdapter(base_path=smf_dir)

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        res = adapter.updater().update_row(
            sheet="Pianificazione",
            id_column="ID ordine",
            id_value="ROW-1",
            updates={
                # assegno stringhe in colonne che possono essere lette come float per via dei NaN
                "Cliente": "Cliente Test",
                "Note": "nota",
            },
        )

    assert res.get("ok") is True
    # Verifica: nessun FutureWarning sul dtype incompatibile
    fw = [w for w in caught if issubclass(w.category, FutureWarning)]
    assert not fw, f"Unexpected FutureWarning(s): {[str(w.message) for w in fw]}"

