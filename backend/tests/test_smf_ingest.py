from pathlib import Path

import openpyxl

from app.ingest.ocr_ingest import (
    normalize_extracted_order,
    write_extracted_order_to_smf,
    write_extracted_orders_to_smf,
)
from app.ingest.ocr_parser import parse_ocr_order_rows
from app.smf.smf_adapter import SMFAdapter


def _build_test_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    codes_ws = wb.active
    codes_ws.title = "Codici"
    codes_ws.append(["Codice", "Descrizione"])
    codes_ws.append(["COD-NEW", "Codice append"])
    codes_ws.append(["COD-UPD", "Codice update"])
    codes_ws.append(["B-UPD", "Codice batch update"])
    codes_ws.append(["B-NEW", "Codice batch new"])

    stations_ws = wb.create_sheet("Postazioni")
    stations_ws.append(["Postazione", "Note"])
    stations_ws.append(["LINEA0", "legacy"])
    stations_ws.append(["LINEA1", "append"])
    stations_ws.append(["LINEA7", "batch update"])
    stations_ws.append(["LINEA8", "batch append"])
    stations_ws.append(["LINEA9", "update"])

    ws = wb.create_sheet("Pianificazione")
    ws.append(
        [
            "ID ordine",
            "Cliente",
            "Codice",
            "Q.ta",
            "Data richiesta cliente",
            "Priorità",
            "Postazione assegnata",
            "Stato (da fare/in corso/finito)",
            "Progress %",
            "Semaforo scadenza",
            "Note",
            "order_id",
            "cliente",
            "codice",
            "qta",
            "postazione",
            "stato",
        ]
    )
    ws.append(
        [
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "EXIST-LOWER",
            "Legacy",
            "OLD",
            1,
            "LINEA0",
            "da fare",
        ]
    )

    log_ws = wb.create_sheet("DiscrepanzeLog")
    log_ws.append(
        [
            "Timestamp",
            "Sorgente",
            "Voce",
            "Valore atteso",
            "Valore trovato",
            "Azione eseguita",
            "Esito",
        ]
    )
    wb.save(path)


def test_normalize_extracted_order_soft_validation():
    result = normalize_extracted_order(
        {
            "order_id": "ORD-001",
            "cliente": "Cliente Test",
            "codice": "COD-001",
            "qta": "12",
            "due_date": "12/04/2026",
            "priority": "urgent",
            "source_file": "scan-001.json",
            "source_type": "ocr_json",
        }
    )

    normalized = result["normalized"]
    assert normalized["order_id"] == "ORD-001"
    assert normalized["qta"] == 12
    assert normalized["due_date"] == "2026-04-12"
    assert normalized["priority"] == "ALTA"
    assert result["discrepancies"]


def test_normalize_extracted_order_accepts_parser_envelope():
    parsed = parse_ocr_order_rows(
        """
        ID ordine: TL-2026-001
        Cliente: Officine Rossi
        Codice: ABC123
        Q.tà: 12
        Data richiesta cliente: 2026-04-20
        Priorità: urgente
        """
    )

    result = normalize_extracted_order(parsed)
    normalized = result["normalized"]
    assert normalized["order_id"] == "TL-2026-001"
    assert normalized["cliente"] == "Officine Rossi"
    assert normalized["codice"] == "ABC123"
    assert normalized["qta"] == 12
    assert normalized["priority"] == "ALTA"


def test_write_extracted_order_to_smf_append_and_update(tmp_path: Path):
    smf_dir = tmp_path / "smf"
    smf_dir.mkdir()
    workbook_path = smf_dir / "SuperMegaFile_Master.xlsx"
    _build_test_workbook(workbook_path)

    adapter = SMFAdapter(base_path=smf_dir)

    append_result = write_extracted_order_to_smf(
        {
            "order_id": "ORD-NEW-001",
            "cliente": "Cliente Nuovo",
            "codice": "COD-NEW",
            "qta": 5,
            "due_date": "2026-04-12",
            "priority": "ALTA",
            "postazione": "LINEA1",
            "note": "append test",
            "source_file": "file-a.json",
            "source_type": "ocr_json",
        },
        adapter=adapter,
    )
    assert append_result["ok"] is True
    assert append_result["write_mode"] == "append_order"
    assert append_result["code_validation"]["status"] == "found"
    assert append_result["station_validation"]["status"] == "found"
    assert append_result["smf_write"]["mode"] == "append_order"
    assert append_result["smf_write"]["result_type"] == "update_not_found_fallback_append"
    assert "ID ordine" in append_result["smf_write"]["written_columns"]
    assert append_result["smf_write"]["fallback_from"]["error"] == "row not found"

    update_result = write_extracted_order_to_smf(
        {
            "order_id": "EXIST-LOWER",
            "cliente": "Legacy Updated",
            "codice": "COD-UPD",
            "qta": 9,
            "postazione": "LINEA9",
            "note": "update test",
            "source_type": "ocr_json",
        },
        adapter=adapter,
    )
    assert update_result["ok"] is True
    assert update_result["write_mode"] == "update_order"
    assert update_result["code_validation"]["status"] == "found"
    assert update_result["station_validation"]["status"] == "found"
    assert update_result["smf_write"]["mode"] == "update_order"
    assert update_result["smf_write"]["result_type"] == "update_order_succeeded"
    assert update_result["smf_write"]["matched_column"] == "order_id"
    assert "Cliente" in update_result["smf_write"]["written_columns"]

    append_without_id_result = write_extracted_order_to_smf(
        {
            "cliente": "Senza ID",
            "codice": "COD-NO-ID",
            "qta": 2,
            "postazione": "LINEA5",
            "source_type": "ocr_json",
        },
        adapter=adapter,
    )
    assert append_without_id_result["ok"] is True
    assert append_without_id_result["write_mode"] == "append_order"
    assert append_without_id_result["station_validation"]["status"] == "missing"
    assert append_without_id_result["smf_write"]["mode"] == "append_order"
    assert append_without_id_result["smf_write"]["result_type"] == "append_without_order_id"
    assert append_without_id_result["smf_write"]["matched_column"] is None

    wb = openpyxl.load_workbook(workbook_path)
    planning = wb["Pianificazione"]
    rows = list(planning.iter_rows(values_only=True))
    assert any("ORD-NEW-001" in row for row in rows)

    matched = [row for row in rows if "EXIST-LOWER" in row]
    assert matched
    row = matched[0]
    assert "Legacy Updated" in row
    assert "COD-UPD" in row
    assert any("COD-NO-ID" in row for row in rows)

    discrepancy_log = wb["DiscrepanzeLog"]
    assert discrepancy_log.max_row >= 2


def test_write_extracted_orders_to_smf_batch(tmp_path: Path):
    smf_dir = tmp_path / "smf_batch"
    smf_dir.mkdir()
    workbook_path = smf_dir / "SuperMegaFile_Master.xlsx"
    _build_test_workbook(workbook_path)

    adapter = SMFAdapter(base_path=smf_dir)

    result = write_extracted_orders_to_smf(
        [
            {
                "order_id": "EXIST-LOWER",
                "cliente": "Batch Updated",
                "codice": "B-UPD",
                "qta": 11,
                "postazione": "LINEA7",
                "source_type": "ocr_json",
            },
            {
                "order_id": "BATCH-NEW-001",
                "cliente": "Batch New",
                "codice": "B-NEW",
                "qta": 3,
                "postazione": "LINEA8",
                "source_type": "ocr_json",
            },
            {
                "cliente": "Batch No Id",
                "codice": "B-NO-ID",
                "qta": 1,
                "source_type": "ocr_json",
            },
        ],
        adapter=adapter,
    )

    assert result["ok"] is True
    assert result["summary"]["total"] == 3
    assert result["summary"]["written"] == 3
    assert result["summary"]["updated"] == 1
    assert result["summary"]["appended"] == 2
    assert result["summary"]["duplicates"] == 0
    assert result["summary"]["warnings"] >= 1
    assert result["summary"]["errors"] == 0
    assert len(result["items"]) == 3
    assert result["items"][0]["write_mode"] == "update_order"
    assert result["items"][1]["write_mode"] == "append_order"


def test_write_extracted_orders_to_smf_batch_deduplicates_order_id(tmp_path: Path):
    smf_dir = tmp_path / "smf_batch_dup"
    smf_dir.mkdir()
    workbook_path = smf_dir / "SuperMegaFile_Master.xlsx"
    _build_test_workbook(workbook_path)

    adapter = SMFAdapter(base_path=smf_dir)

    result = write_extracted_orders_to_smf(
        [
            {
                "order_id": "DUP-001",
                "cliente": "First",
                "codice": "COD-NEW",
                "qta": 1,
                "postazione": "LINEA1",
                "source_type": "ocr_json",
            },
            {
                "order_id": "DUP-001",
                "cliente": "Second",
                "codice": "COD-NEW",
                "qta": 2,
                "postazione": "LINEA1",
                "source_type": "ocr_json",
            },
        ],
        adapter=adapter,
    )

    assert result["ok"] is True
    assert result["summary"]["total"] == 2
    assert result["summary"]["written"] == 1
    assert result["summary"]["duplicates"] == 1
    assert result["items"][1]["write_mode"] == "skipped_duplicate_order_id"
